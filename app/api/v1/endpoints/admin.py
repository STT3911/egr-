"""Admin panel endpoints."""

from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import io
import json
import re
import time
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from fastapi import APIRouter, Depends, File, HTTPException, Request, Response, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session
from starlette.status import HTTP_401_UNAUTHORIZED

from app.core.config import settings
from app.core.database import get_db
from app.core.logger import get_logger
from app.database.models import TradeRegistryImportRun
from app.services.trade_registry_import import empty_trade_registry_stats, infer_source_date

router = APIRouter()
logger = get_logger("api.admin")

ADMIN_COOKIE_NAME = "egr_admin_session"
MAX_IMPORT_ROWS = 5000
TRADE_REGISTRY_UPLOAD_DIR = Path("data/imports/trade_registry/uploads")


def _session_secret() -> str:
    secret = settings.SECRET_KEY or settings.ADMIN_PASSWORD_HASH or settings.ADMIN_PASSWORD
    if not secret:
        secret = "change-me-admin-session-secret"
    return secret


def _b64encode(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode("ascii").rstrip("=")


def _b64decode(data: str) -> bytes:
    padding = "=" * (-len(data) % 4)
    return base64.urlsafe_b64decode(data + padding)


def _sign(payload: str) -> str:
    return hmac.new(_session_secret().encode("utf-8"), payload.encode("ascii"), hashlib.sha256).hexdigest()


def _create_session(username: str) -> str:
    payload = {
        "sub": username,
        "exp": int(time.time()) + settings.ADMIN_SESSION_TTL_HOURS * 3600,
    }
    body = _b64encode(json.dumps(payload, separators=(",", ":")).encode("utf-8"))
    return f"{body}.{_sign(body)}"


def _read_session(token: str | None) -> dict:
    if not token or "." not in token:
        raise HTTPException(status_code=HTTP_401_UNAUTHORIZED, detail="Admin login required")

    body, signature = token.rsplit(".", 1)
    if not hmac.compare_digest(signature, _sign(body)):
        raise HTTPException(status_code=HTTP_401_UNAUTHORIZED, detail="Invalid admin session")

    try:
        payload = json.loads(_b64decode(body))
    except Exception as exc:
        raise HTTPException(status_code=HTTP_401_UNAUTHORIZED, detail="Invalid admin session") from exc

    if int(payload.get("exp", 0)) < int(time.time()):
        raise HTTPException(status_code=HTTP_401_UNAUTHORIZED, detail="Admin session expired")

    return payload


def _iso_datetime(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def _serialize_trade_registry_import(run: TradeRegistryImportRun) -> dict:
    return {
        "id": str(run.id),
        "status": run.status,
        "original_filename": run.original_filename,
        "source_date": run.source_date.isoformat() if run.source_date else None,
        "stats": run.stats_json or {},
        "error": run.error,
        "celery_task_id": run.celery_task_id,
        "created_by": run.created_by,
        "created_at": _iso_datetime(run.created_at),
        "started_at": _iso_datetime(run.started_at),
        "finished_at": _iso_datetime(run.finished_at),
        "updated_at": _iso_datetime(run.updated_at),
    }


def _table_exists(db: Session, table_name: str) -> bool:
    return db.execute(text("SELECT to_regclass(:table_name)"), {"table_name": table_name}).scalar() is not None


def _first_mapping(db: Session, sql: str, params: dict | None = None) -> dict:
    row = db.execute(text(sql), params or {}).mappings().first()
    return dict(row) if row else {}


def _serialize_source_status(
    *,
    key: str,
    name: str,
    updated_at: datetime | None = None,
    source_date: date | datetime | None = None,
    status: str | None = None,
    records_count: int | None = None,
    details: dict | None = None,
    error: str | None = None,
) -> dict:
    return {
        "key": key,
        "name": name,
        "updated_at": _iso_datetime(updated_at),
        "source_date": _iso(source_date),
        "status": status or ("success" if updated_at or source_date else "unknown"),
        "records_count": int(records_count or 0),
        "details": details or {},
        "error": error,
    }


def _latest_run_source(
    db: Session,
    *,
    key: str,
    name: str,
    table_name: str,
    sql: str,
    data_table: str | None = None,
    data_updated_column: str = "last_seen_at",
    records_count_column: str = "*",
    source_date_column: str | None = None,
    details_keys: tuple[str, ...] = (),
) -> dict:
    if not _table_exists(db, table_name):
        return _serialize_source_status(key=key, name=name)

    run = _first_mapping(db, sql)
    data = {}
    if data_table and _table_exists(db, data_table):
        source_date_select = f", max({source_date_column}) AS source_date" if source_date_column else ""
        data = _first_mapping(
            db,
            f"""
            SELECT max({data_updated_column}) AS data_updated_at,
                   count({records_count_column}) AS records_count
                   {source_date_select}
            FROM {data_table}
            """,
        )

    finished_at = run.get("finished_at")
    started_at = run.get("started_at")
    data_updated_at = data.get("data_updated_at")
    updated_at = finished_at or data_updated_at or started_at
    details = {key: run.get(key) for key in details_keys if run.get(key) is not None}

    return _serialize_source_status(
        key=key,
        name=name,
        updated_at=updated_at,
        source_date=data.get("source_date") or run.get("source_date"),
        status=run.get("status"),
        records_count=data.get("records_count") or run.get("records_count") or 0,
        details=details,
        error=run.get("error"),
    )


def _table_snapshot_source(
    db: Session,
    *,
    key: str,
    name: str,
    table_name: str,
    updated_column: str,
    records_count_column: str = "*",
    source_date_column: str | None = None,
) -> dict:
    if not _table_exists(db, table_name):
        return _serialize_source_status(key=key, name=name)

    source_date_select = f", max({source_date_column}) AS source_date" if source_date_column else ""
    row = _first_mapping(
        db,
        f"""
        SELECT max({updated_column}) AS updated_at,
               count({records_count_column}) AS records_count
               {source_date_select}
        FROM {table_name}
        """,
    )
    return _serialize_source_status(
        key=key,
        name=name,
        updated_at=row.get("updated_at"),
        source_date=row.get("source_date"),
        records_count=row.get("records_count") or 0,
    )


def _egr_source(db: Session) -> dict:
    if not _table_exists(db, "egr_companies"):
        return _serialize_source_status(key="egr", name="ЕГР: карточки компаний")

    companies = _first_mapping(
        db,
        """
        SELECT max(updated_at) AS updated_at,
               count(unp) AS records_count
        FROM egr_companies
        """,
    )

    details: dict[str, str | int | None] = {}
    if _table_exists(db, "egr_raw_company_data"):
        raw = _first_mapping(
            db,
            """
            SELECT max(updated_at) AS raw_updated_at,
                   count(unp) AS raw_records_count
            FROM egr_raw_company_data
            """,
        )
        details["raw_updated_at"] = _iso_datetime(raw.get("raw_updated_at"))
        details["raw_records_count"] = int(raw.get("raw_records_count") or 0)

    if _table_exists(db, "egr_system_state"):
        state = _first_mapping(
            db,
            """
            SELECT value AS last_sync_date
            FROM egr_system_state
            WHERE key = 'egr_last_sync_date'
            LIMIT 1
            """,
        )
        details["last_sync_date"] = state.get("last_sync_date")

    return _serialize_source_status(
        key="egr",
        name="ЕГР: карточки компаний",
        updated_at=companies.get("updated_at"),
        records_count=companies.get("records_count") or 0,
        details=details,
    )


def _safe_upload_name(filename: str | None) -> str:
    name = Path(filename or "").name.strip()
    return name or "trade_registry.csv"


async def _save_uploaded_file(file: UploadFile, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("wb") as handle:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            handle.write(chunk)


def require_admin(request: Request) -> dict:
    return _read_session(request.cookies.get(ADMIN_COOKIE_NAME))


def _verify_password(password: str) -> bool:
    if settings.ADMIN_PASSWORD_HASH:
        digest = hashlib.sha256(password.encode("utf-8")).hexdigest()
        return hmac.compare_digest(digest, settings.ADMIN_PASSWORD_HASH)

    if settings.ADMIN_PASSWORD:
        return hmac.compare_digest(password, settings.ADMIN_PASSWORD)

    return hmac.compare_digest(password, "admin")


def _is_admin_password_configured() -> bool:
    return bool(settings.ADMIN_PASSWORD_HASH or settings.ADMIN_PASSWORD)


def _cookie_secure() -> bool:
    if settings.ADMIN_COOKIE_SECURE is not None:
        return settings.ADMIN_COOKIE_SECURE
    return settings.APP_ENV == "production"


def _set_admin_cookie(response: Response, token: str) -> None:
    response.set_cookie(
        ADMIN_COOKIE_NAME,
        token,
        max_age=settings.ADMIN_SESSION_TTL_HOURS * 3600,
        httponly=True,
        secure=_cookie_secure(),
        samesite="lax",
        path="/",
    )


def _clear_admin_cookie(response: Response) -> None:
    response.delete_cookie(ADMIN_COOKIE_NAME, path="/")


def _normalize_unp(value: object) -> str | None:
    if value is None:
        return None
    text_value = str(value).strip()
    if not text_value:
        return None
    if text_value.endswith(".0"):
        text_value = text_value[:-2]
    digits = re.sub(r"\D", "", text_value)
    return digits if re.fullmatch(r"\d{9}", digits) else None


def _extract_unps_from_rows(rows: Iterable[list[object]]) -> list[str]:
    materialized = [[cell for cell in row] for row in rows]
    if not materialized:
        return []

    header = [str(cell or "").strip().lower() for cell in materialized[0]]
    unp_index = next(
        (idx for idx, value in enumerate(header) if value in {"unp", "унп", "у н п"}),
        None,
    )
    data_rows = materialized[1:] if unp_index is not None else materialized

    unps: list[str] = []
    for row in data_rows:
        candidates = [row[unp_index]] if unp_index is not None and unp_index < len(row) else row[:1]
        unp = next((_normalize_unp(value) for value in candidates if _normalize_unp(value)), None)
        if unp:
            unps.append(unp)

    return unps


async def _read_uploaded_unps(file: UploadFile) -> list[str]:
    contents = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        sheet = workbook.active
        rows = ([cell for cell in row] for row in sheet.iter_rows(values_only=True))
        return _extract_unps_from_rows(rows)

    if filename.endswith(".csv"):
        text_contents = None
        for encoding in ("utf-8-sig", "cp1251", "utf-8"):
            try:
                text_contents = contents.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text_contents is None:
            raise HTTPException(status_code=400, detail="CSV encoding is not supported")

        sample = text_contents[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text_contents), dialect)
        return _extract_unps_from_rows(reader)

    raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are supported")


EXCEL_CELL_LIMIT = 32767


def _iso(value: date | datetime | None) -> str | None:
    return value.isoformat() if value else None


def _cell_value(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT:
            return value[: EXCEL_CELL_LIMIT - 14] + "...[truncated]"
        return value
    return str(value)


def _append_sheet(workbook: Workbook, title: str, headers: list[str], rows: Iterable[dict]) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    sheet.append(headers)
    for row in rows:
        sheet.append([_cell_value(row.get(header)) for header in headers])

    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 60)


def _append_export_sheet(
    workbook: Workbook,
    title: str,
    columns: list[tuple[str, str]],
    rows: Iterable[dict],
) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    sheet.append([label for _, label in columns])
    for row in rows:
        sheet.append([_cell_value(row.get(key)) for key, _ in columns])

    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 60)


def _fetch_rows(db: Session, sql: str, unps: list[str], label: str = "query") -> list[dict]:
    if not unps:
        return []
    started_at = time.perf_counter()
    rows = db.execute(text(sql), {"unps": [int(unp) for unp in unps]}).mappings().all()
    elapsed = time.perf_counter() - started_at
    logger.info("Admin export %s: %s rows in %.3fs", label, len(rows), elapsed)
    return [dict(row) for row in rows]


def _get_company_rows(db: Session, unps: list[str]) -> dict[str, list[dict]]:
    if not unps:
        return {}

    export: dict[str, list[dict]] = {}
    export["Summary"] = _fetch_rows(
        db,
        """
        WITH requested(unp) AS (
            SELECT unnest(CAST(:unps AS bigint[]))
        )
        SELECT
            r.unp,
            CASE WHEN c.unp IS NULL THEN 'Not found' ELSE 'Found' END AS result,
            c.current_status_code,
            rs.name AS current_status_name,
            c.registration_date,
            c.liquidation_date,
            cm.name AS creation_method,
            ca.name AS creation_authority,
            cha.name AS current_authority,
            lm.name AS liquidation_reason,
            la.name AS liquidation_authority,
            et.name AS entity_type,
            n.full_name_ru AS current_full_name_ru,
            n.short_name_ru AS current_short_name_ru,
            n.full_name_by AS current_full_name_by,
            COALESCE(pl.address, a.full_address) AS current_address,
            pl.fetched_at AS place_location_fetched_at,
            cv.ved AS current_ved,
            cc.email,
            cc.phone,
            cc.website
        FROM requested r
        LEFT JOIN egr_companies c ON c.unp = r.unp
        LEFT JOIN LATERAL (
            SELECT full_name_ru, short_name_ru, full_name_by
            FROM egr_company_names_history
            WHERE company_id = c.id
            ORDER BY (valid_to IS NULL) DESC, valid_to DESC NULLS LAST, valid_from DESC NULLS LAST
            LIMIT 1
        ) n ON true
        LEFT JOIN LATERAL (
            SELECT full_address
            FROM egr_company_addresses_history
            WHERE company_id = c.id
            ORDER BY (valid_to IS NULL) DESC, valid_to DESC NULLS LAST, valid_from DESC NULLS LAST
            LIMIT 1
        ) a ON true
        LEFT JOIN egr_company_place_locations pl ON pl.unp = c.unp
        LEFT JOIN LATERAL (
            SELECT email, phone, website
            FROM egr_company_contacts_history
            WHERE company_id = c.id
            ORDER BY (valid_to IS NULL) DESC, valid_to DESC NULLS LAST, valid_from DESC NULLS LAST
            LIMIT 1
        ) cc ON true
        LEFT JOIN LATERAL (
            SELECT string_agg(concat_ws(' ', ved_code, ved_name), '; ' ORDER BY ved_code) AS ved
            FROM egr_company_ved_history
            WHERE company_id = c.id AND valid_to IS NULL
        ) cv ON true
        LEFT JOIN ref_statuses rs ON rs.id = c.current_status_code
        LEFT JOIN ref_creation_methods cm ON cm.id = c.creation_method_id
        LEFT JOIN ref_authorities ca ON ca.id = c.creation_authority_id
        LEFT JOIN ref_authorities cha ON cha.id = c.current_authority_id
        LEFT JOIN ref_liquidation_methods lm ON lm.id = c.liquidation_reason_id
        LEFT JOIN ref_authorities la ON la.id = c.liquidation_authority_id
        LEFT JOIN ref_entity_types et ON et.id = c.entity_type_id
        ORDER BY array_position(CAST(:unps AS bigint[]), r.unp)
        """,
        unps,
        "Summary",
    )

    export["Names"] = _fetch_rows(db, """
        SELECT c.unp, n.full_name_ru, n.short_name_ru, n.full_name_by, n.search_name, n.valid_from, n.valid_to
        FROM egr_company_names_history n
        JOIN egr_companies c ON c.id = n.company_id
        WHERE c.unp = ANY(CAST(:unps AS bigint[]))
        ORDER BY c.unp, n.valid_from DESC NULLS LAST, n.valid_to DESC NULLS LAST
    """, unps, "Names")

    export["Addresses"] = _fetch_rows(db, """
        SELECT c.unp, a.full_address, a.postal_code, a.region, a.district, a.valid_from, a.valid_to
        FROM egr_company_addresses_history a
        JOIN egr_companies c ON c.id = a.company_id
        WHERE c.unp = ANY(CAST(:unps AS bigint[]))
        ORDER BY c.unp, a.valid_from DESC NULLS LAST, a.valid_to DESC NULLS LAST
    """, unps, "Addresses")

    export["VED"] = _fetch_rows(db, """
        SELECT c.unp, v.ved_code, v.ved_name, v.valid_from, v.valid_to
        FROM egr_company_ved_history v
        JOIN egr_companies c ON c.id = v.company_id
        WHERE c.unp = ANY(CAST(:unps AS bigint[]))
        ORDER BY c.unp, v.valid_from DESC NULLS LAST, v.ved_code
    """, unps, "VED")

    export["Contacts"] = _fetch_rows(db, """
        SELECT c.unp, ct.email, ct.website, ct.phone, ct.valid_from, ct.valid_to
        FROM egr_company_contacts_history ct
        JOIN egr_companies c ON c.id = ct.company_id
        WHERE c.unp = ANY(CAST(:unps AS bigint[]))
        ORDER BY c.unp, ct.valid_from DESC NULLS LAST, ct.valid_to DESC NULLS LAST
    """, unps, "Contacts")

    export["Events"] = _fetch_rows(db, """
        SELECT
            c.unp,
            e.event_record_id,
            re.name AS event_type,
            e.event_date,
            e.cancel_date,
            e.document_date,
            e.deadline_date,
            e.suspension_end_date,
            e.document_number,
            da.name AS decision_authority,
            doa.name AS document_authority,
            rf.name AS foundation,
            e.notes
        FROM egr_company_events e
        JOIN egr_companies c ON c.id = e.company_id
        LEFT JOIN ref_events re ON re.id = e.event_type_id
        LEFT JOIN ref_authorities da ON da.id = e.decision_authority_id
        LEFT JOIN ref_authorities doa ON doa.id = e.document_authority_id
        LEFT JOIN ref_foundations rf ON rf.id = e.foundation_id
        WHERE c.unp = ANY(CAST(:unps AS bigint[]))
        ORDER BY c.unp, e.event_date DESC NULLS LAST, e.created_at DESC NULLS LAST
    """, unps, "Events")

    export["GRP"] = _fetch_rows(db, """
        SELECT unp, full_name, short_name, registration_date, inspectorate_code, inspectorate_name,
               status_code, status_date, address, fetched_at
        FROM grp_taxpayer_data
        WHERE unp = ANY(CAST(:unps AS bigint[]))
        ORDER BY unp
    """, unps, "GRP")

    export["Tax Debt"] = _fetch_rows(db, """
        SELECT debtor_unp AS unp, imns_code, imns_name, debt_date, repayment_date, slice_date
        FROM nalog_debt_records
        WHERE debtor_unp = ANY(CAST(:unps AS bigint[]))
        ORDER BY debtor_unp, slice_date DESC, imns_code
    """, unps, "Tax Debt")

    export["GIAS Accreditation"] = _fetch_rows(db, """
        SELECT unp, name, uid_customer, customer_id, summary, state, dt_update, dt_from, dt_to,
               is_customer, is_organizator, phone, email, web_site, region, city_name,
               placements_address, placements_country, placements_post_index, placements_city,
               placements_address_detail, okogu_name, okogu_code
        FROM gias_accredited_customers
        WHERE unp IN (SELECT unnest(CAST(:unps AS bigint[]))::text)
        ORDER BY unp
    """, unps, "GIAS Accreditation")

    export["Locked Suppliers"] = _fetch_rows(db, """
        SELECT
            ls.provider_unp AS unp,
            ls.name,
            ls.state,
            ls.location,
            ls.reg_number,
            ls.add_date,
            ls.del_date,
            ai.initials AS author_initials,
            bi.text AS base_incl_text,
            be.text AS base_excl_text
        FROM locked_suppliers ls
        LEFT JOIN locked_supplier_authors ai ON ai.id = ls.author_id
        LEFT JOIN locked_supplier_reasons bi ON bi.id = ls.base_incl_id
        LEFT JOIN locked_supplier_reasons be ON be.id = ls.base_excl_id
        WHERE ls.provider_unp IN (SELECT unnest(CAST(:unps AS bigint[]))::text)
        ORDER BY ls.provider_unp, ls.add_date DESC NULLS LAST
    """, unps, "Locked Suppliers")

    export["Trade Registry"] = _fetch_rows(db, """
        SELECT
            unp, legal_name, legal_address, object_type, object_name, internet_shop_domain,
            trade_network_name, object_region, object_district, object_locality, object_street,
            object_building, object_office, object_contacts, format_type, location_type,
            assortment_type, is_firm, trade_object_type, trade_area, retail_trade, wholesale_trade,
            registration_number, inclusion_date, source_date, source_file
        FROM trade_registry_records
        WHERE unp = ANY(CAST(:unps AS bigint[]))
        ORDER BY unp, source_date DESC NULLS LAST, registration_number
    """, unps, "Trade Registry")

    export["Licenses"] = _fetch_rows(db, """
        SELECT
            holder_unp AS unp,
            generated_number,
            holder_name,
            activity_type_name,
            activity_date_start,
            activity_date_end,
            activity_is_active,
            last_seen_at
        FROM license_records
        WHERE holder_unp = ANY(CAST(:unps AS bigint[]))
        ORDER BY holder_unp, activity_is_active DESC NULLS LAST, last_seen_at DESC NULLS LAST, license_id DESC
    """, unps, "Licenses")

    return export


def _build_result_workbook(unps: list[str], company_rows: dict[str, list[dict]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheet_columns = {
        "Summary": [
            ("unp", "УНП"),
            ("result", "Результат"),
            ("current_status_code", "Код текущего статуса"),
            ("current_status_name", "Текущий статус"),
            ("registration_date", "Дата регистрации"),
            ("liquidation_date", "Дата ликвидации"),
            ("creation_method", "Способ создания"),
            ("creation_authority", "Орган создания"),
            ("current_authority", "Текущий регистрирующий орган"),
            ("liquidation_reason", "Причина ликвидации"),
            ("liquidation_authority", "Орган ликвидации"),
            ("entity_type", "Тип субъекта"),
            ("current_full_name_ru", "Полное наименование RU"),
            ("current_short_name_ru", "Краткое наименование RU"),
            ("current_full_name_by", "Полное наименование BY"),
            ("current_address", "Текущий адрес"),
            ("place_location_fetched_at", "Дата получения адреса"),
            ("current_ved", "Текущие ВЭД"),
            ("email", "Email"),
            ("phone", "Телефон"),
            ("website", "Сайт"),
        ],
        "Names": [
            ("unp", "УНП"),
            ("full_name_ru", "Полное наименование RU"),
            ("short_name_ru", "Краткое наименование RU"),
            ("full_name_by", "Полное наименование BY"),
            ("search_name", "Поисковое наименование"),
            ("valid_from", "Действует с"),
            ("valid_to", "Действует по"),
        ],
        "Addresses": [
            ("unp", "УНП"),
            ("full_address", "Адрес"),
            ("postal_code", "Почтовый индекс"),
            ("region", "Область"),
            ("district", "Район"),
            ("valid_from", "Действует с"),
            ("valid_to", "Действует по"),
        ],
        "VED": [
            ("unp", "УНП"),
            ("ved_code", "Код ВЭД"),
            ("ved_name", "Наименование ВЭД"),
            ("valid_from", "Действует с"),
            ("valid_to", "Действует по"),
        ],
        "Contacts": [
            ("unp", "УНП"),
            ("email", "Email"),
            ("website", "Сайт"),
            ("phone", "Телефон"),
            ("valid_from", "Действует с"),
            ("valid_to", "Действует по"),
        ],
        "Events": [
            ("unp", "УНП"),
            ("event_record_id", "ID события"),
            ("event_type", "Тип события"),
            ("event_date", "Дата события"),
            ("cancel_date", "Дата отмены"),
            ("document_date", "Дата документа"),
            ("deadline_date", "Срок"),
            ("suspension_end_date", "Окончание приостановления"),
            ("document_number", "Номер документа"),
            ("decision_authority", "Орган решения"),
            ("document_authority", "Орган документа"),
            ("foundation", "Основание"),
            ("notes", "Примечание"),
        ],
        "GRP": [
            ("unp", "УНП"),
            ("full_name", "Полное наименование"),
            ("short_name", "Краткое наименование"),
            ("registration_date", "Дата регистрации"),
            ("inspectorate_code", "Код инспекции"),
            ("inspectorate_name", "Инспекция"),
            ("status_code", "Код статуса"),
            ("status_date", "Дата статуса"),
            ("address", "Адрес"),
            ("fetched_at", "Дата получения"),
        ],
        "Tax Debt": [
            ("unp", "УНП"),
            ("imns_code", "Код ИМНС"),
            ("imns_name", "ИМНС"),
            ("debt_date", "Дата задолженности"),
            ("repayment_date", "Дата погашения"),
            ("slice_date", "Дата среза"),
        ],
        "GIAS Accreditation": [
            ("unp", "УНП"),
            ("name", "Наименование"),
            ("uid_customer", "UID заказчика"),
            ("customer_id", "ID заказчика"),
            ("summary", "Описание"),
            ("state", "Статус"),
            ("dt_update", "Дата обновления"),
            ("dt_from", "Действует с"),
            ("dt_to", "Действует по"),
            ("is_customer", "Заказчик"),
            ("is_organizator", "Организатор"),
            ("phone", "Телефон"),
            ("email", "Email"),
            ("web_site", "Сайт"),
            ("region", "Регион"),
            ("city_name", "Город"),
            ("placements_address", "Адрес размещения"),
            ("placements_country", "Страна размещения"),
            ("placements_post_index", "Почтовый индекс"),
            ("placements_city", "Город размещения"),
            ("placements_address_detail", "Детали адреса"),
            ("okogu_name", "ОКОГУ"),
            ("okogu_code", "Код ОКОГУ"),
        ],
        "Locked Suppliers": [
            ("unp", "УНП"),
            ("name", "Наименование"),
            ("state", "Статус"),
            ("location", "Местонахождение"),
            ("reg_number", "Регистрационный номер"),
            ("add_date", "Дата включения"),
            ("del_date", "Дата исключения"),
            ("author_initials", "Автор"),
            ("base_incl_text", "Основание включения"),
            ("base_excl_text", "Основание исключения"),
        ],
        "Trade Registry": [
            ("unp", "УНП"),
            ("legal_name", "Юридическое наименование"),
            ("legal_address", "Юридический адрес"),
            ("object_type", "Тип объекта"),
            ("object_name", "Наименование объекта"),
            ("internet_shop_domain", "Домен интернет-магазина"),
            ("trade_network_name", "Торговая сеть"),
            ("object_region", "Область объекта"),
            ("object_district", "Район объекта"),
            ("object_locality", "Населенный пункт"),
            ("object_street", "Улица"),
            ("object_building", "Дом"),
            ("object_office", "Помещение"),
            ("object_contacts", "Контакты объекта"),
            ("format_type", "Формат"),
            ("location_type", "Тип расположения"),
            ("assortment_type", "Ассортимент"),
            ("is_firm", "Фирменный"),
            ("trade_object_type", "Тип торгового объекта"),
            ("trade_area", "Торговая площадь"),
            ("retail_trade", "Розничная торговля"),
            ("wholesale_trade", "Оптовая торговля"),
            ("registration_number", "Регистрационный номер"),
            ("inclusion_date", "Дата включения"),
            ("source_date", "Дата источника"),
            ("source_file", "Файл источника"),
        ],
        "Licenses": [
            ("unp", "UNP"),
            ("generated_number", "License number"),
            ("holder_name", "License holder"),
            ("activity_type_name", "Activity type"),
            ("activity_date_start", "Start date"),
            ("activity_date_end", "End date"),
            ("activity_is_active", "Active"),
            ("last_seen_at", "Checked at"),
        ],
    }

    for title, columns in sheet_columns.items():
        _append_export_sheet(workbook, title, columns, company_rows.get(title, []))

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@router.post("/login")
async def login(payload: dict, response: Response):
    if settings.APP_ENV == "production" and not _is_admin_password_configured():
        raise HTTPException(status_code=500, detail="Admin password is not configured")

    username = str(payload.get("username", "")).strip()
    password = str(payload.get("password", ""))

    if not hmac.compare_digest(username, settings.ADMIN_USERNAME) or not _verify_password(password):
        raise HTTPException(status_code=HTTP_401_UNAUTHORIZED, detail="Invalid login or password")

    _set_admin_cookie(response, _create_session(username))
    return {"username": username}


@router.post("/logout")
async def logout(response: Response):
    _clear_admin_cookie(response)
    return {"ok": True}


@router.get("/me")
async def me(session: dict = Depends(require_admin)):
    return {"username": session["sub"]}


@router.get("/data-sources")
async def list_data_sources(
    db: Session = Depends(get_db),
    session: dict = Depends(require_admin),
):
    sources = [
        _egr_source(db),
        _table_snapshot_source(
            db,
            key="grp",
            name="ГРП: реестр плательщиков",
            table_name="grp_taxpayer_data",
            updated_column="fetched_at",
            records_count_column="unp",
        ),
        _table_snapshot_source(
            db,
            key="place_locations",
            name="ЕГР Mobile: адреса",
            table_name="egr_company_place_locations",
            updated_column="fetched_at",
            records_count_column="unp",
        ),
        _latest_run_source(
            db,
            key="trade_registry",
            name="МАРТ: торговый реестр",
            table_name="trade_registry_import_runs",
            sql="""
                SELECT status,
                       started_at,
                       finished_at,
                       source_date,
                       error,
                       stats_json ->> 'rows' AS rows,
                       stats_json ->> 'written' AS written,
                       stats_json ->> 'matched' AS matched
                FROM trade_registry_import_runs
                ORDER BY created_at DESC
                LIMIT 1
            """,
            data_table="trade_registry_records",
            data_updated_column="last_seen_at",
            records_count_column="id",
            source_date_column="source_date",
            details_keys=("rows", "written", "matched"),
        ),
        _latest_run_source(
            db,
            key="licenses",
            name="license.gov.by: лицензии",
            table_name="license_sync_runs",
            sql="""
                SELECT status,
                       started_at,
                       finished_at,
                       error,
                       processed_records AS records_count,
                       created_count,
                       updated_count,
                       unchanged_count,
                       failed_count
                FROM license_sync_runs
                ORDER BY started_at DESC
                LIMIT 1
            """,
            data_table="license_records",
            data_updated_column="last_seen_at",
            records_count_column="id",
            details_keys=("created_count", "updated_count", "unchanged_count", "failed_count"),
        ),
        _latest_run_source(
            db,
            key="gias_accredited_customers",
            name="ГИАС: аккредитованные заказчики",
            table_name="gias_sync_runs",
            sql="""
                SELECT status,
                       started_at,
                       finished_at,
                       error,
                       records_fetched AS records_count,
                       created_count,
                       updated_count,
                       unchanged_count,
                       history_created_count
                FROM gias_sync_runs
                WHERE registry_name = 'accredited_customers'
                ORDER BY started_at DESC
                LIMIT 1
            """,
            data_table="gias_accredited_customers",
            data_updated_column="last_seen_at",
            records_count_column="id",
            details_keys=("created_count", "updated_count", "unchanged_count", "history_created_count"),
        ),
        _latest_run_source(
            db,
            key="gias_locked_suppliers",
            name="ГИАС: недобросовестные поставщики",
            table_name="gias_sync_runs",
            sql="""
                SELECT status,
                       started_at,
                       finished_at,
                       error,
                       records_fetched AS records_count,
                       created_count,
                       updated_count,
                       unchanged_count,
                       history_created_count
                FROM gias_sync_runs
                WHERE registry_name = 'locked_suppliers'
                ORDER BY started_at DESC
                LIMIT 1
            """,
            data_table="locked_suppliers",
            data_updated_column="last_seen_at",
            records_count_column="id",
            details_keys=("created_count", "updated_count", "unchanged_count", "history_created_count"),
        ),
        _table_snapshot_source(
            db,
            key="pvt_residents",
            name="ПВТ: резиденты",
            table_name="pvt_resident_records",
            updated_column="last_seen_at",
            records_count_column="id",
        ),
        _table_snapshot_source(
            db,
            key="eaeu_sez_residents",
            name="ЕАЭС: резиденты СЭЗ",
            table_name="eaeu_sez_resident_records",
            updated_column="last_seen_at",
            records_count_column="id",
        ),
        _table_snapshot_source(
            db,
            key="nalog_debt",
            name="МНС: задолженности",
            table_name="nalog_debt_records",
            updated_column="created_at",
            records_count_column="id",
            source_date_column="slice_date",
        ),
        _latest_run_source(
            db,
            key="bankrot",
            name="bankrot.gov.by: дела о банкротстве",
            table_name="bankrot_sync_runs",
            sql="""
                SELECT status,
                       started_at,
                       finished_at,
                       error,
                       processed_cases AS records_count,
                       total_cases,
                       failed_cases,
                       last_page
                FROM bankrot_sync_runs
                ORDER BY started_at DESC
                LIMIT 1
            """,
            data_table="bankrot_cases",
            data_updated_column="updated_at",
            records_count_column="case_id",
            details_keys=("total_cases", "failed_cases", "last_page"),
        ),
    ]
    return {
        "updated_at": datetime.utcnow().isoformat(),
        "items": sources,
    }


@router.get("/companies")
async def list_companies(
    q: str = "",
    limit: int = 25,
    offset: int = 0,
    db: Session = Depends(get_db),
    session: dict = Depends(require_admin),
):
    limit = min(max(limit, 1), 100)
    offset = max(offset, 0)
    cleaned = re.sub(r"[\s-]", "", q.strip())

    where = ""
    params: dict[str, object] = {"limit": limit, "offset": offset}
    if cleaned:
        if cleaned.isdigit():
            where = "WHERE CAST(c.unp AS TEXT) LIKE :q || '%'"
            params["q"] = cleaned
        else:
            where = "WHERE n.full_name_ru ILIKE :pattern OR n.short_name_ru ILIKE :pattern"
            params["pattern"] = f"%{q.strip()}%"

    sql = text(
        f"""
        WITH current_names AS (
            SELECT DISTINCT ON (company_id)
                company_id,
                full_name_ru,
                short_name_ru
            FROM egr_company_names_history
            ORDER BY company_id, (valid_to IS NULL) DESC, valid_to DESC NULLS LAST, valid_from DESC NULLS LAST
        ),
        current_addresses AS (
            SELECT DISTINCT ON (company_id)
                company_id,
                full_address
            FROM egr_company_addresses_history
            ORDER BY company_id, (valid_to IS NULL) DESC, valid_to DESC NULLS LAST, valid_from DESC NULLS LAST
        )
        SELECT
            c.unp,
            n.full_name_ru,
            n.short_name_ru,
            COALESCE(pl.address, a.full_address) AS address,
            rs.name AS status_name,
            c.registration_date
        FROM egr_companies c
        LEFT JOIN current_names n ON n.company_id = c.id
        LEFT JOIN current_addresses a ON a.company_id = c.id
        LEFT JOIN egr_company_place_locations pl ON pl.unp = c.unp
        LEFT JOIN ref_statuses rs ON rs.id = c.current_status_code
        {where}
        ORDER BY c.updated_at DESC NULLS LAST, c.unp DESC
        LIMIT :limit OFFSET :offset
        """
    )
    rows = db.execute(sql, params).mappings().all()

    count_sql = text(
        f"""
        WITH current_names AS (
            SELECT DISTINCT ON (company_id)
                company_id,
                full_name_ru,
                short_name_ru
            FROM egr_company_names_history
            ORDER BY company_id, (valid_to IS NULL) DESC, valid_to DESC NULLS LAST, valid_from DESC NULLS LAST
        )
        SELECT count(*) AS total
        FROM egr_companies c
        LEFT JOIN current_names n ON n.company_id = c.id
        {where}
        """
    )
    total = db.execute(count_sql, params).scalar() or 0

    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "items": [
            {
                "unp": int(row["unp"]),
                "name": row["full_name_ru"] or row["short_name_ru"],
                "short_name": row["short_name_ru"],
                "address": row["address"],
                "status": row["status_name"],
                "registration_date": _iso(row["registration_date"]),
            }
            for row in rows
        ],
    }


@router.post("/trade-registry/imports")
async def create_trade_registry_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    session: dict = Depends(require_admin),
):
    filename = _safe_upload_name(file.filename)
    if not filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are supported")

    run_id = uuid.uuid4()
    stored_path = TRADE_REGISTRY_UPLOAD_DIR / f"{run_id}_{filename}"
    source_date = infer_source_date(Path(filename)) or date.today()

    try:
        await _save_uploaded_file(file, stored_path)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to save uploaded file: {exc}") from exc

    stats = empty_trade_registry_stats()
    stats["source_date"] = source_date.isoformat()
    run = TradeRegistryImportRun(
        id=run_id,
        status="queued",
        original_filename=filename,
        stored_path=str(stored_path),
        source_date=source_date,
        stats_json=stats,
        created_by=session.get("sub"),
    )
    db.add(run)
    db.commit()
    db.refresh(run)

    try:
        from app.tasks.trade_registry_tasks import import_trade_registry_csv_task

        async_result = import_trade_registry_csv_task.delay(str(run.id))
        run.celery_task_id = async_result.id
        db.add(run)
        db.commit()
        db.refresh(run)
    except Exception as exc:
        run.status = "failed"
        run.error = f"Failed to queue Celery task: {exc}"
        run.finished_at = datetime.utcnow()
        db.add(run)
        db.commit()
        raise HTTPException(status_code=503, detail=run.error) from exc

    return _serialize_trade_registry_import(run)


@router.get("/trade-registry/imports")
async def list_trade_registry_imports(
    limit: int = 10,
    db: Session = Depends(get_db),
    session: dict = Depends(require_admin),
):
    safe_limit = max(1, min(limit, 50))
    runs = (
        db.query(TradeRegistryImportRun)
        .order_by(TradeRegistryImportRun.created_at.desc())
        .limit(safe_limit)
        .all()
    )
    return {"items": [_serialize_trade_registry_import(run) for run in runs]}


@router.get("/trade-registry/imports/{run_id}")
async def get_trade_registry_import(
    run_id: uuid.UUID,
    db: Session = Depends(get_db),
    session: dict = Depends(require_admin),
):
    run = db.get(TradeRegistryImportRun, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Trade registry import was not found")
    return _serialize_trade_registry_import(run)


@router.post("/pvt-residents/sync")
async def queue_pvt_residents_sync(
    limit: int | None = None,
    offset: int = 0,
    start_unp: int | None = None,
    batch_size: int = 100,
    delay: float = 0.2,
    only_missing: bool = False,
    resume: bool = True,
    source: str = "catalog",
    output: str | None = None,
    letters: str | None = None,
    prefixes: str | None = None,
    session: dict = Depends(require_admin),
):
    if limit is not None and limit < 1:
        raise HTTPException(status_code=400, detail="limit must be greater than zero")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be greater than or equal to zero")
    if start_unp is not None and start_unp < 1:
        raise HTTPException(status_code=400, detail="start_unp must be greater than zero")
    if batch_size < 1:
        raise HTTPException(status_code=400, detail="batch_size must be greater than zero")
    if delay < 0:
        raise HTTPException(status_code=400, detail="delay must be greater than or equal to zero")
    if source not in {"catalog", "egr_scan"}:
        raise HTTPException(status_code=400, detail="source must be either catalog or egr_scan")

    from app.tasks.park_tasks import sync_pvt_residents_task

    task = sync_pvt_residents_task.delay(
        limit=limit,
        offset=offset,
        start_unp=start_unp,
        batch_size=batch_size,
        delay=delay,
        only_missing=only_missing,
        resume=resume,
        source=source,
        output=output,
        letters=letters,
        prefixes=prefixes,
    )
    return {
        "queued": True,
        "task_id": task.id,
        "limit": limit,
        "offset": offset,
        "start_unp": start_unp,
        "batch_size": batch_size,
        "delay": delay,
        "only_missing": only_missing,
        "resume": resume,
        "source": source,
        "output": output,
        "letters": letters,
        "prefixes": prefixes,
    }


@router.post("/fill-company-file")
async def fill_company_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    session: dict = Depends(require_admin),
):
    unps = await _read_uploaded_unps(file)
    if not unps:
        raise HTTPException(status_code=400, detail="No valid 9-digit UNP values found")
    if len(unps) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=400, detail=f"Too many rows. Limit is {MAX_IMPORT_ROWS}")

    company_rows = _get_company_rows(db, unps)
    result = _build_result_workbook(unps, company_rows)
    output_name = "company-fill-result.xlsx"

    return StreamingResponse(
        io.BytesIO(result),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{output_name}"'},
    )
