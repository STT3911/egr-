"""Excel dossier export for a single company."""
from __future__ import annotations

import io
import json
from datetime import datetime, timezone
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.crud.company import CompanyCRUD
from app.database.models import CompanyEvent, GrpTaxpayerData, NalogDebtRecord
from app.services.company_relations import find_related_by_address, find_related_by_contact


EXCEL_CELL_LIMIT = 32_000
JSON_CHUNK_SIZE = 30_000
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="EAF2F8")
LIGHT_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))


def _value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool, datetime)):
        text = value.isoformat() if isinstance(value, datetime) else value
    else:
        text = json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(text, str) and text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    if isinstance(text, str) and len(text) > EXCEL_CELL_LIMIT:
        return text[: EXCEL_CELL_LIMIT - 16] + "… [сокращено]"
    return text


def _json_chunks(value: Any) -> list[str]:
    text = json.dumps(value, ensure_ascii=False, default=str)
    return [text[index : index + JSON_CHUNK_SIZE] for index in range(0, len(text), JSON_CHUNK_SIZE)] or [""]


def _add_sheet(
    workbook: Workbook,
    title: str,
    columns: list[tuple[str, str]],
    rows: Iterable[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    sheet.append([label for _, label in columns])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_count = 0
    for row in rows:
        sheet.append([_value(row.get(key)) for key, _ in columns])
        row_count += 1

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(row_count + 1, 1)}"
    sheet.sheet_view.showGridLines = False
    for column_index, (_, label) in enumerate(columns, start=1):
        values = [str(label)]
        values.extend(str(sheet.cell(row=row_index, column=column_index).value or "") for row_index in range(2, min(row_count + 2, 102)))
        width = min(max(max(map(len, values)) + 2, 12), 60)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _event_rows(db: Session, company_id: Any) -> list[dict[str, Any]]:
    events = (
        db.query(CompanyEvent)
        .filter(CompanyEvent.company_id == company_id)
        .order_by(CompanyEvent.event_date.desc().nullslast(), CompanyEvent.created_at.desc())
        .all()
    )
    return [
        {
            "event_record_id": item.event_record_id,
            "event_type": item.event_type.name if item.event_type else None,
            "event_date": item.event_date,
            "cancel_date": item.cancel_date,
            "document_date": item.document_date,
            "deadline_date": item.deadline_date,
            "suspension_end_date": item.suspension_end_date,
            "document_number": item.document_number,
            "decision_authority": item.decision_authority.name if item.decision_authority else None,
            "document_authority": item.document_authority.name if item.document_authority else None,
            "foundation": item.foundation.name if item.foundation else None,
            "notes": item.notes,
        }
        for item in events
    ]


def _grp_rows(db: Session, unp: int) -> list[dict[str, Any]]:
    item = db.query(GrpTaxpayerData).filter(GrpTaxpayerData.unp == unp).first()
    if not item:
        return []
    return [
        {
            "full_name": item.full_name,
            "short_name": item.short_name,
            "registration_date": item.registration_date,
            "inspectorate_code": item.inspectorate_code,
            "inspectorate_name": item.inspectorate_name,
            "status_code": item.status_code,
            "status_date": item.status_date,
            "address": item.address,
            "fetched_at": item.fetched_at,
            "updated_at": item.updated_at,
        }
    ]


def _tax_debt_rows(db: Session, unp: int) -> list[dict[str, Any]]:
    items = (
        db.query(NalogDebtRecord)
        .filter(NalogDebtRecord.debtor_unp == unp)
        .order_by(NalogDebtRecord.slice_date.desc(), NalogDebtRecord.debt_date.desc().nullslast())
        .all()
    )
    return [
        {
            "imns_code": item.imns_code,
            "imns_name": item.imns_name,
            "debt_date": item.debt_date,
            "repayment_date": item.repayment_date,
            "slice_date": item.slice_date,
        }
        for item in items
    ]


def _display_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    return text[:10] if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-" else text


def _compact_lines(rows: Iterable[dict[str, Any]], formatter: Any) -> str:
    values: list[str] = []
    seen: set[str] = set()
    for row in rows:
        value = str(formatter(row) or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        values.append(value)
    if not values:
        return "Нет данных"
    return "\n".join(f"{index}. {value}" for index, value in enumerate(values, start=1))


def _period_for_rows(
    rows: Iterable[dict[str, Any]],
    start_keys: tuple[str, ...] = (),
    end_keys: tuple[str, ...] = (),
    observed_keys: tuple[str, ...] = (),
) -> str:
    row_list = list(rows)
    starts = sorted(
        value
        for row in row_list
        for key in start_keys
        if (value := _display_date(row.get(key)))
    )
    ends = sorted(
        value
        for row in row_list
        for key in end_keys
        if (value := _display_date(row.get(key)))
    )
    if starts or ends:
        start = starts[0] if starts else "начало не указано"
        has_open_period = bool(start_keys and end_keys) and any(
            any(row.get(key) not in (None, "") for key in start_keys)
            and not any(row.get(key) not in (None, "") for key in end_keys)
            for row in row_list
        )
        end = "по настоящее время" if has_open_period else (ends[-1] if ends else starts[-1])
        return f"{start} — {end}"

    observed = sorted(
        value
        for row in row_list
        for key in observed_keys
        if (value := _display_date(row.get(key)))
    )
    if observed:
        return f"Состояние на {observed[-1]}" if observed[0] == observed[-1] else f"{observed[0]} — {observed[-1]}"
    return "Период в источнике не указан"


def _add_overview_sheet(
    workbook: Workbook,
    unp: int,
    company_name: str,
    status: str | None,
    generated_at: datetime,
    company_period: str,
    rows: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(title="Сводка")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A7"

    sheet.merge_cells("A1:E1")
    title = sheet["A1"]
    title.value = f"Досье компании {company_name or unp}"
    title.fill = HEADER_FILL
    title.font = Font(color="FFFFFF", bold=True, size=16)
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet["A2"] = "УНП"
    sheet["B2"] = unp
    sheet["C2"] = "Статус"
    sheet["D2"] = status or "—"
    sheet.merge_cells("D2:E2")
    sheet["A3"] = "Период компании"
    sheet["B3"] = company_period
    sheet.merge_cells("B3:C3")
    sheet["D3"] = "Сформировано UTC"
    sheet["E3"] = generated_at.replace(microsecond=0).isoformat()
    for cell in (sheet["A2"], sheet["C2"], sheet["A3"], sheet["D3"]):
        cell.font = Font(bold=True, color="1E3A5F")

    headers = ["Раздел", "Количество", "Данные", "Период данных", "Полные данные"]
    sheet.append([])
    sheet.append([])
    sheet.append(headers)
    for cell in sheet[6]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[6].height = 28

    for item in rows:
        sheet.append([
            item["section"],
            item["count"],
            _value(item["data"]),
            item["period"],
            item["sheet"],
        ])
        row_number = sheet.max_row
        sheet.cell(row=row_number, column=1).fill = SECTION_FILL
        sheet.cell(row=row_number, column=1).font = Font(bold=True, color="1E3A5F")
        link_cell = sheet.cell(row=row_number, column=5)
        link_cell.hyperlink = f"#'{item['sheet']}'!A1"
        link_cell.style = "Hyperlink"
        for cell in sheet[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = LIGHT_BORDER
        sheet.cell(row=row_number, column=2).alignment = Alignment(horizontal="center", vertical="top")

    sheet.auto_filter.ref = f"A6:E{sheet.max_row}"
    sheet.column_dimensions["A"].width = 27
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 72
    sheet.column_dimensions["D"].width = 28
    sheet.column_dimensions["E"].width = 24
    for row_number in range(7, sheet.max_row + 1):
        line_count = str(sheet.cell(row=row_number, column=3).value or "").count("\n") + 1
        sheet.row_dimensions[row_number].height = min(max(30, line_count * 15), 120)


def build_company_report(db: Session, unp: int) -> bytes | None:
    company_crud = CompanyCRUD(db)
    profile = company_crud.get_full_dossier(unp)
    company = company_crud.get_by_unp(unp)
    if not profile or not company:
        return None

    bankruptcy = company_crud.get_bankrot_dossier(unp)
    grp_rows = _grp_rows(db, unp)
    tax_rows = _tax_debt_rows(db, unp)
    related_by_contact = find_related_by_contact(db, company.id, limit=200)
    related_by_address = find_related_by_address(db, company.id, limit=200)
    events = _event_rows(db, company.id)

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = f"Досье компании {unp}"
    workbook.properties.subject = "Агрегированные данные по компании"
    workbook.properties.creator = "TENDERS.BY"

    generated_at = datetime.now(timezone.utc)
    company_name = profile.get("current_name_ru") or profile.get("current_short_name_ru") or ""
    names = profile.get("names") or []
    addresses = profile.get("addresses") or []
    ved_rows = profile.get("ved") or []
    contacts = profile.get("contacts_aggregated") or profile.get("contacts") or []
    leadership_rows = profile.get("leadership_observations") or []
    trade_rows = profile.get("trade_registry_records") or []
    license_rows = profile.get("license_records") or []
    inspection_rows = profile.get("inspection_plan_records") or []
    certificates = profile.get("belltpp_own_certificates") or []
    bankrot_cases = bankruptcy.get("cases") or []
    pvt = profile.get("pvt_resident")
    accreditation = profile.get("gias_accreditation")
    locked_suppliers = profile.get("gias_locked_suppliers") or []
    sez_rows = profile.get("eaeu_sez_resident_records") or []
    related_rows = [
        {**item, "relation_type": "Совпадение контакта"} for item in related_by_contact
    ] + [
        {**item, "relation_type": "Совпадение адреса"} for item in related_by_address
    ]

    company_period = _period_for_rows(
        [{"start": profile.get("registration_date"), "end": profile.get("liquidation_date")}],
        ("start",),
        ("end",),
    )
    overview_rows = [
        {
            "section": "Основные сведения",
            "count": 1,
            "data": "\n".join(filter(None, [
                company_name,
                f"Статус: {profile.get('current_status_name') or '—'}",
                f"Адрес: {profile.get('place_location_address') or '—'}",
            ])),
            "period": company_period,
            "sheet": "Названия ЕГР",
        },
        {
            "section": "Названия ЕГР",
            "count": len(names),
            "data": _compact_lines(names, lambda row: row.get("full_name_ru") or row.get("short_name_ru") or row.get("full_name_by")),
            "period": _period_for_rows(names, ("valid_from",), ("valid_to",)),
            "sheet": "Названия ЕГР",
        },
        {
            "section": "Адреса ЕГР",
            "count": len(addresses),
            "data": _compact_lines(addresses, lambda row: row.get("full_address")),
            "period": _period_for_rows(addresses, ("valid_from",), ("valid_to",)),
            "sheet": "Адреса ЕГР",
        },
        {
            "section": "Виды деятельности",
            "count": len(ved_rows),
            "data": _compact_lines(ved_rows, lambda row: " — ".join(filter(None, [str(row.get("ved_code") or ""), str(row.get("ved_name") or "")]))),
            "period": _period_for_rows(ved_rows, ("valid_from",), ("valid_to",)),
            "sheet": "ВЭД ЕГР",
        },
        {
            "section": "Контакты",
            "count": len(contacts),
            "data": _compact_lines(contacts, lambda row: f"{row.get('contact_type') or 'контакт'}: {row.get('value') or row.get('email') or row.get('phone') or row.get('website') or '—'}"),
            "period": _period_for_rows(contacts, observed_keys=("last_seen_at", "updated_at")),
            "sheet": "Контакты",
        },
        {
            "section": "Руководители в публикациях",
            "count": len(leadership_rows),
            "data": _compact_lines(
                leadership_rows,
                lambda row: " — ".join(
                    filter(
                        None,
                        [
                            str(row.get("person_name") or ""),
                            str(row.get("position") or ""),
                            _display_date(row.get("event_date")),
                        ],
                    )
                ),
            ),
            "period": _period_for_rows(leadership_rows, observed_keys=("event_date",)),
            "sheet": "Руководители публикации",
        },
        {
            "section": "События ЕГР",
            "count": len(events),
            "data": _compact_lines(events, lambda row: " — ".join(filter(None, [_display_date(row.get("event_date")), str(row.get("event_type") or ""), str(row.get("document_number") or "")]))),
            "period": _period_for_rows(events, observed_keys=("event_date", "document_date", "created_at")),
            "sheet": "События ЕГР",
        },
        {
            "section": "Задолженность МНС",
            "count": len(tax_rows),
            "data": _compact_lines(tax_rows, lambda row: f"{row.get('imns_name') or 'ИМНС'}: долг с {_display_date(row.get('debt_date')) or 'дата не указана'}, погашение {_display_date(row.get('repayment_date')) or 'не указано'}"),
            "period": _period_for_rows(tax_rows, ("debt_date",), ("repayment_date",), ("slice_date",)),
            "sheet": "Задолженность МНС",
        },
        {
            "section": "Дела о банкротстве",
            "count": len(bankrot_cases),
            "data": _compact_lines(bankrot_cases, lambda row: " — ".join(filter(None, [str(row.get("number") or row.get("case_id") or ""), str(row.get("status") or ""), str(row.get("procedure_type") or "")]))),
            "period": _period_for_rows(bankrot_cases, ("start_date",), ("end_date",), ("updated_at",)),
            "sheet": "Дела о банкротстве",
        },
        {
            "section": "Торговый реестр МАРТ",
            "count": len(trade_rows),
            "data": _compact_lines(trade_rows, lambda row: " — ".join(filter(None, [str(row.get("object_type") or ""), str(row.get("object_name") or row.get("internet_shop_domain") or ""), str(row.get("object_locality") or "")]))),
            "period": _period_for_rows(trade_rows, observed_keys=("inclusion_date", "source_date", "last_seen_at")),
            "sheet": "Торговый реестр МАРТ",
        },
        {
            "section": "ПВТ",
            "count": 1 if pvt else 0,
            "data": _compact_lines([pvt] if pvt else [], lambda row: " — ".join(filter(None, [str(row.get("name") or ""), str(row.get("city") or ""), str(row.get("website") or "")]))),
            "period": _period_for_rows([pvt] if pvt else [], observed_keys=("last_seen_at",)),
            "sheet": "ПВТ",
        },
        {
            "section": "Лицензии",
            "count": len(license_rows),
            "data": _compact_lines(license_rows, lambda row: " — ".join(filter(None, [str(row.get("generated_number") or ""), str(row.get("activity_type_name") or ""), "активна" if row.get("activity_is_active") else "неактивна"]))),
            "period": _period_for_rows(license_rows, ("activity_date_start",), ("activity_date_end",), ("last_seen_at",)),
            "sheet": "Лицензии",
        },
        {
            "section": "Планы проверок",
            "count": len(inspection_rows),
            "data": _compact_lines(inspection_rows, lambda row: " — ".join(filter(None, [str(row.get("plan_period") or ""), str(row.get("controller_authority") or ""), str(row.get("start_month") or "")]))),
            "period": _period_for_rows(inspection_rows, observed_keys=("plan_period", "last_seen_at")),
            "sheet": "Планы проверок",
        },
        {
            "section": "Связанные компании",
            "count": len(related_rows),
            "data": _compact_lines(related_rows, lambda row: f"{row.get('unp') or '—'} — {row.get('name') or 'Без названия'} ({row.get('relation_type')})"),
            "period": f"Расчёт на {generated_at.date().isoformat()}",
            "sheet": "Связи по контактам" if related_by_contact else "Связи по адресу",
        },
    ]
    _add_overview_sheet(
        workbook,
        unp,
        company_name,
        profile.get("current_status_name"),
        generated_at,
        company_period,
        overview_rows,
    )

    _add_sheet(workbook, "Названия ЕГР", [("full_name_ru", "Полное RU"), ("short_name_ru", "Краткое RU"), ("full_name_by", "Полное BY"), ("valid_from", "С"), ("valid_to", "По")], names)
    _add_sheet(workbook, "Адреса ЕГР", [("full_address", "Адрес"), ("postal_code", "Индекс"), ("region", "Область"), ("district", "Район"), ("valid_from", "С"), ("valid_to", "По")], addresses)
    _add_sheet(workbook, "ВЭД ЕГР", [("ved_code", "Код"), ("ved_name", "Наименование"), ("valid_from", "С"), ("valid_to", "По")], ved_rows)
    _add_sheet(workbook, "Контакты", [("contact_type", "Тип"), ("value", "Значение"), ("full_name", "Контактное лицо"), ("position", "Должность"), ("sources", "Источники"), ("email", "Email"), ("phone", "Телефон"), ("website", "Сайт"), ("fax", "Факс")], contacts)
    _add_sheet(workbook, "Руководители публикации", [("person_name", "ФИО"), ("position", "Должность"), ("organization_name", "Организация в источнике"), ("event_date", "Дата списка"), ("exam_type", "Вид проверки"), ("source_url", "Источник"), ("match_method", "Метод привязки"), ("match_confidence", "Уверенность")], leadership_rows)
    _add_sheet(workbook, "События ЕГР", [("event_record_id", "ID"), ("event_type", "Тип"), ("event_date", "Дата"), ("cancel_date", "Отмена"), ("document_date", "Документ"), ("deadline_date", "Срок"), ("suspension_end_date", "Окончание приостановления"), ("document_number", "Номер документа"), ("decision_authority", "Орган решения"), ("document_authority", "Орган документа"), ("foundation", "Основание"), ("notes", "Примечание")], events)
    _add_sheet(workbook, "ГРП МНС", [("full_name", "Полное название"), ("short_name", "Краткое название"), ("registration_date", "Регистрация"), ("inspectorate_code", "Код инспекции"), ("inspectorate_name", "Инспекция"), ("status_code", "Статус"), ("status_date", "Дата статуса"), ("address", "Адрес"), ("fetched_at", "Получено"), ("updated_at", "Обновлено")], grp_rows)
    _add_sheet(workbook, "Задолженность МНС", [("imns_code", "Код ИМНС"), ("imns_name", "ИМНС"), ("debt_date", "Дата долга"), ("repayment_date", "Погашение"), ("slice_date", "Дата среза")], tax_rows)

    _add_sheet(workbook, "Связи по контактам", [("unp", "УНП"), ("name", "Наименование"), ("matched_type", "Тип"), ("matched_value", "Совпадение")], related_by_contact)
    _add_sheet(workbook, "Связи по адресу", [("unp", "УНП"), ("name", "Наименование"), ("address", "Адрес")], related_by_address)

    _add_sheet(workbook, "ПВТ", [("name", "Наименование"), ("city", "Город"), ("legal_address", "Адрес"), ("phone", "Телефон"), ("website", "Сайт"), ("activity_directions", "Направления"), ("description", "Описание"), ("profile_url", "Профиль"), ("last_seen_at", "Проверено")], [pvt] if pvt else [])
    _add_sheet(workbook, "Торговый реестр МАРТ", [("registration_number", "Номер"), ("legal_name", "Юр. название"), ("legal_address", "Юр. адрес"), ("object_type", "Тип объекта"), ("object_name", "Объект"), ("internet_shop_domain", "Интернет-магазин"), ("trade_network_name", "Сеть"), ("object_region", "Область"), ("object_locality", "Населённый пункт"), ("object_street", "Улица"), ("object_building", "Дом"), ("object_office", "Помещение"), ("object_contacts", "Контакты"), ("goods_groups", "Группы товаров"), ("inclusion_date", "Включено"), ("source_date", "Дата источника")], trade_rows)
    _add_sheet(workbook, "ГИАС аккредитация", [("state", "Статус"), ("summary", "Описание"), ("phone", "Телефон"), ("email", "Email"), ("web_site", "Сайт"), ("city_name", "Город"), ("placements_address", "Адрес"), ("dt_from", "С"), ("dt_to", "По"), ("dt_update", "Обновлено")], [accreditation] if accreditation else [])
    _add_sheet(workbook, "ГИАС поставщики", [("state", "Статус"), ("name", "Наименование"), ("location", "Место"), ("reg_number", "Номер"), ("add_date", "Включён"), ("del_date", "Исключён"), ("base_incl_text", "Основание включения"), ("base_excl_text", "Основание исключения"), ("author_initials", "Автор")], locked_suppliers)
    _add_sheet(workbook, "ЕАЭС и СЭЗ", [("country", "Страна"), ("full_name", "Название"), ("legal_address", "Адрес"), ("registration_agency", "Орган регистрации"), ("sez_name", "СЭЗ"), ("project_name", "Проект"), ("registry_entry_date", "Дата включения"), ("certificate", "Свидетельство"), ("source_url", "Источник")], sez_rows)
    _add_sheet(workbook, "Лицензии", [("generated_number", "Номер"), ("holder_name", "Владелец"), ("activity_type_name", "Вид деятельности"), ("activity_date_start", "С"), ("activity_date_end", "По"), ("activity_is_active", "Активна"), ("last_seen_at", "Проверено")], license_rows)
    _add_sheet(workbook, "Планы проверок", [("plan_period", "Период"), ("source_region", "Регион"), ("plan_title", "План"), ("plan_item_no", "Пункт"), ("approving_authority", "Утвердивший орган"), ("controller_unp", "УНП контролёра"), ("controller_authority", "Контролирующий орган"), ("executor_phone", "Телефон"), ("start_month", "Месяц"), ("source_file", "Источник")], inspection_rows)

    _add_sheet(workbook, "Сертификаты БелТПП", [("holder_name", "Владелец"), ("cert_number", "Номер"), ("blank_number", "Бланк"), ("issue_date", "Выдан"), ("valid_until", "Действует до"), ("verify_url", "Проверка"), ("last_seen_at", "Проверено")], certificates)
    product_rows = []
    for certificate in certificates:
        for product in certificate.get("products") or []:
            product_rows.append({"cert_number": certificate.get("cert_number"), **product})
    _add_sheet(workbook, "Продукция БелТПП", [("cert_number", "Сертификат"), ("row_no", "Строка"), ("name", "Продукция"), ("code", "Код")], product_rows)

    _add_sheet(workbook, "Дела о банкротстве", [("case_id", "ID"), ("number", "Номер"), ("start_date", "Начало"), ("end_date", "Окончание"), ("status", "Статус"), ("procedure_type", "Процедура"), ("court", "Суд"), ("judge", "Судья"), ("manager_id", "ID управляющего"), ("manager_name", "Управляющий"), ("last_judgment_id", "Последнее решение"), ("fetch_error", "Ошибка"), ("updated_at", "Обновлено")], bankrot_cases)
    dataset_rows = []
    for case in bankrot_cases:
        for dataset in case.get("datasets") or []:
            chunks = _json_chunks(dataset.get("payload"))
            for chunk_number, payload_chunk in enumerate(chunks, start=1):
                dataset_rows.append(
                    {
                        "case_id": case.get("case_id"),
                        "case_number": case.get("number"),
                        "dataset_type": dataset.get("dataset_type"),
                        "endpoint": dataset.get("endpoint"),
                        "http_method": dataset.get("http_method"),
                        "chunk_number": chunk_number,
                        "chunks_total": len(chunks),
                        "payload": payload_chunk,
                        "fetch_error": dataset.get("fetch_error"),
                        "fetched_at": dataset.get("fetched_at"),
                    }
                )
    _add_sheet(workbook, "Bankrot данные", [("case_id", "ID дела"), ("case_number", "Номер дела"), ("dataset_type", "Раздел"), ("endpoint", "Endpoint"), ("http_method", "Метод"), ("chunk_number", "Часть"), ("chunks_total", "Всего частей"), ("payload", "JSON payload"), ("fetch_error", "Ошибка"), ("fetched_at", "Получено")], dataset_rows)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
