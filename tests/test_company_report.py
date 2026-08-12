from io import BytesIO
from types import SimpleNamespace
from unittest.mock import Mock, patch

from openpyxl import load_workbook

from app.services import company_report


def test_excel_dossier_does_not_expose_risk_scoring() -> None:
    crud = Mock()
    crud.get_full_dossier.return_value = {
        "current_name_ru": "Тестовая компания",
        "current_status_name": "Действующий",
    }
    crud.get_by_unp.return_value = SimpleNamespace(id=1)
    crud.get_bankrot_dossier.return_value = {}

    with (
        patch.object(company_report, "CompanyCRUD", return_value=crud),
        patch.object(company_report, "_grp_rows", return_value=[]),
        patch.object(company_report, "_tax_debt_rows", return_value=[]),
        patch.object(company_report, "_event_rows", return_value=[]),
        patch.object(company_report, "find_related_by_contact", return_value=[]),
        patch.object(company_report, "find_related_by_address", return_value=[]),
    ):
        report_bytes = company_report.build_company_report(Mock(), 123456789)

    assert report_bytes is not None
    workbook = load_workbook(BytesIO(report_bytes), read_only=True)
    assert "Риск-профиль" not in workbook.sheetnames

    summary_sections = {
        row[0]
        for row in workbook["Сводка"].iter_rows(min_row=7, values_only=True)
        if row and row[0]
    }
    assert "Риск-профиль" not in summary_sections
